
from sys import prefix
from tracemalloc import start
from turtle import width
import matplotlib
import h5py
matplotlib.use('Agg')#仅保存图片不显示
import hdf5storage
from flask import before_render_template
import skvideo.io
import numpy as np
import matplotlib.pyplot as plt
import scipy
import scipy.io as sio
import os
from utils import preprocess_raw_video, bvpsnr, peak_delay,video_duration,cross_corr,read_video
from preprocess import generate_pulse_gt
from mp4_wav import *
from inference_preprocess import detrend
from scipy.signal import butter
from sklearn import preprocessing
from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook
import cv2
from skimage.util import img_as_float
# HDF5_DISABLE_VERSION_CHECK = 1

#experiments---------------TBD
subject_id = 1
experiment_nums = 20

def mp4_to_mat(input_video,start_time):
    # dim1 = 320
    # dim2=240
    # vidObj = cv2.VideoCapture(input_video)
    # totalFrames = int(vidObj.get(cv2.CAP_PROP_FRAME_COUNT)) # get total frame size
    # Xsub = np.zeros((1800, dim1, dim2, 3), dtype = np.float32)
    # height = vidObj.get(cv2.CAP_PROP_FRAME_HEIGHT)
    # width = vidObj.get(cv2.CAP_PROP_FRAME_WIDTH)
    # success, img = vidObj.read()
    # dims = img.shape
    # print(dims)
    # print("Orignal frames", totalFrames)
    # print("Orignal Height", height)
    # print("Original width", width)


    RGB_length,full_video_duration =video_duration(input_video)
    print(RGB_length,full_video_duration)
    RGB = read_video(input_video)

    end_time = start_time+60
    start_frame = int(start_time*RGB_length/full_video_duration)
    end_frame = int(end_time*RGB_length/full_video_duration)
    frames_num =end_frame-start_frame
    print(frames_num)
    dim_width, dim_height = 320,240
    RGB_cut = RGB[start_frame:end_frame]
    for i in range(frames_num):
        RGB_cut[i] = cv2.resize(img_as_float(RGB_cut[i]), (dim_width, dim_height), interpolation = cv2.INTER_AREA)
    # cv2.show(RGB_cut[0])
    # RGB_cut = cv2.resize(img_as_float(RGB_cut[:, int(width/2)-int(height/2 + 1):int(height/2)+int(width/2), :])
    #         , (dim_width, dim_height), interpolation = cv2.INTER_AREA)
    print(RGB_cut.shape())
    return RGB_cut

def audio_sync(win_sound_file,video_sound_file):
    fs = 16000
    fig1 = np.memmap(win_sound_file, dtype='h', mode='r')
    fig2 = np.memmap(video_sound_file, dtype='h', mode='r')
    
    fig1 = np.array(list(fig1)).flatten()
    fig2 = np.array(list(fig2)).flatten()
    # print(fig1,fig2)
    fig_mask1 = np.argwhere((fig1>=-30000)&(fig1<=30000))
    fig1 =np.take(fig1,fig_mask1)
    fig_mask2 = np.argwhere((fig2>=-30000)&(fig2<=30000))
    fig2 =np.take(fig2,fig_mask2)
    scaler = MinMaxScaler(feature_range=(0, 1))  #将数据归一到0到1，可以根据数据特点归一到-1到1
    fig1 = scaler.fit_transform(fig1)  #归一化
    fig2 = scaler.fit_transform(fig2)  #归一化
    fig1 = np.array(list(fig1)).flatten()
    fig2 = np.array(list(fig2)).flatten()
    fig1[0:5]=0
    fig2[0:5]=0
    [b_pulse, a_pulse] = scipy.signal.butter(3, [500 / fs * 2, 3000/ fs * 2], btype='bandpass') #调节滤波频率
    fig1 = scipy.signal.filtfilt(b_pulse, a_pulse, np.double(fig1))
    fig2 = scipy.signal.filtfilt(b_pulse, a_pulse, np.double(fig2))
    fig1_corr,fig2_corr,delay = cross_corr(fig1,fig2)

    delay_time = delay/16000
    return delay_time

def data_to_mat(subject_id):
    
    origin_path =  'E:/rPPG/dataset'#数据集根目录
    current_path = os.path.join(origin_path,str(subject_id))#实验者目录
    os.chdir(current_path) #切换当前路径
    excel_path = str(subject_id)+'.xlsx'
    subjects_excel_path = 'E:/rPPG/experiments/被试信息表.xlsx' #被试信息表
    wb = load_workbook(excel_path)
    ws = wb['实验信息总表']
    subjects_wb = load_workbook(subjects_excel_path)
    subjects_ws = subjects_wb['Sheet1']
    if subjects_ws.cell(subject_id+1,9).value == '男':
        gender = 'male'
    elif subjects_ws.cell(subject_id+1,9).value == '女':
        gender = 'female'
    else:
        print('gender error!')
    skin_color =  subjects_ws.cell(subject_id+1,11).value
    if subjects_ws.cell(subject_id+1,12).value == '有':
        glasses = 'True'
    elif subjects_ws.cell(subject_id+1,12).value == '无':
        glasses = 'False'
    else:
        print('glasses error!')   
    if subjects_ws.cell(subject_id+1,13).value == '有':
        hair_cover = 'True'
    elif subjects_ws.cell(subject_id+1,13).value == '无':
        hair_cover = 'False'
    else:
        print('hair_cover error!')  
    if subjects_ws.cell(subject_id+1,14).value == '有':
        makeup = 'True'
    elif subjects_ws.cell(subject_id+1,14).value == '无':
        makeup = 'False'
    else:
        print('makeup error!')  

    for i in range(8,experiment_nums):
        experiment_id = i
        if experiment_id<16 and experiment_id>=12:
            exercise = 'True'
        else:
            exercise = 'False'
        prefix ='E:/rPPG/mat_dataset/'+ 'subject'+ str(subject_id)
        # if os.path.exists(str(prefix)) is False:
        #         os.makedirs(str(prefix))
        filepath = str(subject_id)+'_'+str(experiment_id)+'.mp4' 
        mat_filepath = prefix+'/p' +str(subject_id)+'_'+str(experiment_id)+'.mat'
        wav_file = str(subject_id)+'_'+str(experiment_id)+'.wav'
        
        wavefile1 = str(subject_id)+'_'+str(experiment_id)+'start.wav'
        wavefile2 = str(subject_id)+'_'+str(experiment_id)+'end.wav'
        RGB_length,full_video_duration =video_duration(filepath)
        print(RGB_length,full_video_duration)
        extract_audio(filepath,wav_file)
        delay_time_1 = audio_sync(wavefile1,wav_file)
        delay_time_2 = audio_sync(wavefile2,wav_file)
        print(delay_time_1,delay_time_2)
        start_win_record_time = ws.cell(experiment_id+2,4).value
        record_delay_time = ws.cell(experiment_id+2,5).value
        light = ws.cell(experiment_id+2,2).value
        motion = ws.cell(experiment_id+2,3).value


        # if i ==12:
        #   delay_time_1 = delay_time_2 - record_delay_time
        relative_delay_time = int(abs(record_delay_time-(delay_time_2-delay_time_1))*1000)
        

        GT_path = excel_path #---------------TBD
        print("working on: ", filepath)
        gt_pulse, gt_time = generate_pulse_gt(GT_path,experiment_id)
        gt_start_time = int(gt_time[0])
        video_duration_time = 60
        gt_end_time = gt_start_time + video_duration_time * 1000

        start_time = delay_time_1+ (gt_start_time-start_win_record_time)/1000
        end_time = start_time + video_duration_time
        # end_time = start_time+time_duration
        ws.cell(2+i,22).value = relative_delay_time
        

        # relative_delay_time = 0 #----------to be deleted


        if abs(relative_delay_time) > 30:
            print("---------------------------------同步异常 : ", relative_delay_time,"---------------------------------同步异常")
        else:
            print("同步正常 : ", relative_delay_time)
            if os.path.exists(str(prefix)) is False:
                os.makedirs(str(prefix))
            RGB_cut,RGB_length = preprocess_raw_video(filepath,start_time)
            for i in range(len(gt_pulse)):
                if gt_time[i]>=gt_end_time:
                    print(i,gt_time[i])
                    gt_pulse = gt_pulse[0:i]
                    gt_time = gt_time[0:i]
                    break
                    
            re_time = np.linspace(gt_start_time,gt_end_time,RGB_length)
            re_gt_pulse = np.interp(re_time, gt_time, gt_pulse)
            zero_pulse = np.zeros(1800-RGB_length)
            re_gt_pulse = np.concatenate((re_gt_pulse,zero_pulse),axis=0)
            gt_length = len(re_gt_pulse)
            print('GT_ppg length:',gt_length)
            data_dict = {'video':RGB_cut,'GT_ppg':re_gt_pulse,'light':light,'motion':motion,
                        'exercise':exercise,'skin_color':skin_color,'gender':gender,'glasser':glasses,'hair_cover':hair_cover,'makeup':makeup}
            sio.savemat(mat_filepath,data_dict,do_compression= True)
            
            # hdf5storage.savemat( mat_filepath, data_dict, format=7.3, matlab_compatible=True, compress=True )
            # np.save(str(prefix)+'/video', RGB )
            # np.save(str(prefix)+'/bvp', re_gt_pulse)

    wb.save(excel_path)  
        # ws.cell(experiment_id+2,5).value = record_delay_time


if __name__ == "__main__":
    subject_id(1)
