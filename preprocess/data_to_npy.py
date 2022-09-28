
from sys import prefix
from tracemalloc import start
import matplotlib
matplotlib.use('Agg')#仅保存图片不显示

from flask import before_render_template
import skvideo.io
import numpy as np
import matplotlib.pyplot as plt
import scipy
import scipy.io as sio
import os
from utils import bottom_delay, bvpsnr, peak_delay,video_duration,cross_corr,read_video
from preprocess import generate_pulse_gt
from mp4_wav import *
from inference_preprocess import detrend
from scipy.signal import butter
from sklearn import preprocessing
from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook
HDF5_DISABLE_VERSION_CHECK = 1

#experiments---------------TBD
subject_id = 35
experiment_nums = 20

def audio_sync(win_sound_file,video_sound_file):
    fs = 16000
    fig1 = np.memmap(win_sound_file, dtype='h', mode='r')
    fig2 = np.memmap(video_sound_file, dtype='h', mode='r')
    
    fig1 = np.array(list(fig1)).flatten()
    fig2 = np.array(list(fig2)).flatten()
    # print(fig1,fig2)
    fig_mask1 = np.argwhere((fig1>=-30000)&(fig1<=30000))
    fig1 =np.take(fig1,fig_mask1)
    fig_mask2 = np.argwhere((fig2>=-30000)&(fig2<=30000))
    fig2 =np.take(fig2,fig_mask2)
    scaler = MinMaxScaler(feature_range=(0, 1))  #将数据归一到0到1，可以根据数据特点归一到-1到1
    fig1 = scaler.fit_transform(fig1)  #归一化
    fig2 = scaler.fit_transform(fig2)  #归一化
    fig1 = np.array(list(fig1)).flatten()
    fig2 = np.array(list(fig2)).flatten()
    fig1[0:5]=0
    fig2[0:5]=0
    [b_pulse, a_pulse] = scipy.signal.butter(3, [500 / fs * 2, 3000/ fs * 2], btype='bandpass') #调节滤波频率
    fig1 = scipy.signal.filtfilt(b_pulse, a_pulse, np.double(fig1))
    fig2 = scipy.signal.filtfilt(b_pulse, a_pulse, np.double(fig2))
    fig1_corr,fig2_corr,delay = cross_corr(fig1,fig2)

    delay_time = delay/16000
    return delay_time

if __name__ == "__main__":
    
    origin_path =  r'E:\rPPG\dataset'#数据集根目录
    current_path = os.path.join(origin_path,str(subject_id))#实验者目录
    os.chdir(current_path) #切换当前路径
    excel_path = str(subject_id)+'.xlsx'
    wb = load_workbook(excel_path)
    ws = wb['实验信息总表']
    

    for i in range(experiment_nums):
        experiment_id = i
        
        prefix ='E:/rPPG/processed_dataset/'+ 'subject'+ str(subject_id)+'_'+str(experiment_id)
        # if os.path.exists(str(prefix)) is False:
        #         os.makedirs(str(prefix))
        filepath = str(subject_id)+'_'+str(experiment_id)+'.mp4' 
        cut_filepath = prefix+'/video.mp4' 
        wav_file = str(subject_id)+'_'+str(experiment_id)+'.wav'
        
        wavefile1 = str(subject_id)+'_'+str(experiment_id)+'start.wav'
        wavefile2 = str(subject_id)+'_'+str(experiment_id)+'end.wav'
        RGB_length,full_video_duration =video_duration(filepath)
        print(RGB_length,full_video_duration)
        extract_audio(filepath,wav_file)
        delay_time_1 = audio_sync(wavefile1,wav_file)
        delay_time_2 = audio_sync(wavefile2,wav_file)
        print(delay_time_1,delay_time_2)
        start_win_record_time = ws.cell(experiment_id+2,4).value
        record_delay_time = ws.cell(experiment_id+2,5).value
        # if i ==12:
        #   delay_time_1 = delay_time_2 - record_delay_time
        relative_delay_time = int(abs(record_delay_time-(delay_time_2-delay_time_1))*1000)
        

        GT_path = excel_path #---------------TBD
        print("working on: ", filepath)
        gt_pulse, gt_time = generate_pulse_gt(GT_path,experiment_id)
        gt_start_time = int(gt_time[0])
        video_duration_time = 60
        gt_end_time = gt_start_time + video_duration_time * 1000

        start_time = delay_time_1+ (gt_start_time-start_win_record_time)/1000
        end_time = start_time + video_duration_time
        # end_time = start_time+time_duration
        start_frame = int(start_time*RGB_length/full_video_duration)
        end_frame = int(end_time*RGB_length/full_video_duration)
        ws.cell(2+i,22).value = relative_delay_time
        

        # relative_delay_time = 0 #----------to be deleted


        if abs(relative_delay_time) > 30:
            print("---------------------------------同步异常 : ", relative_delay_time,"---------------------------------同步异常")
        else:
            print("同步正常 : ", relative_delay_time)
            if os.path.exists(str(prefix)) is False:
                os.makedirs(str(prefix))
            extract_video(filepath,cut_filepath,seconds_to_time(start_time))
            
            RGB_length,full_video_duration =video_duration(cut_filepath)
            
            print(RGB_length,full_video_duration)
            RGB_length = int(RGB_length)
            for i in range(len(gt_pulse)):
                if gt_time[i]>=gt_end_time:
                    print(i,gt_time[i])
                    gt_pulse = gt_pulse[0:i]
                    gt_time = gt_time[0:i]
                    break
                    
            re_time = np.linspace(gt_start_time,gt_end_time,RGB_length)
            re_gt_pulse = np.interp(re_time, gt_time, gt_pulse)
            gt_length = len(re_gt_pulse)
            
            # np.save(str(prefix)+'/video', RGB )
            np.save(str(prefix)+'/bvp', re_gt_pulse)
    wb.save(excel_path)  
        # ws.cell(experiment_id+2,5).value = record_delay_time

