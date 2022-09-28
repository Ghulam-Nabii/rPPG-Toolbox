
import matplotlib
matplotlib.use('Agg')#仅保存图片不显示

from flask import before_render_template
import skvideo.io
import numpy as np
import matplotlib.pyplot as plt
import scipy
import scipy.io as sio
import os
from utils import bottom_delay, bvpsnr, peak_delay,video_duration,cross_corr,read_video
from preprocess import generate_pulse_gt
from mp4_wav import *
from inference_preprocess import detrend
from scipy.signal import butter
from sklearn import preprocessing
from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook
HDF5_DISABLE_VERSION_CHECK = 1

#experiments---------------TBD
subject_id = 14
experiment_nums = 8

def calculate_HR(fs, signal):
    # N = 30 * fs
    N = 60 * fs
    pulse_fft = np.expand_dims(signal, 0)
    f, pxx = scipy.signal.periodogram(pulse_fft, fs=fs, nfft=4 * N, detrend=False)
    fmask = np.argwhere((f >= 0.75) & (f <= 2.5))  # regular Heart beat are 0.75*60 and 2.5*60
    # fmask = np.argwhere(((f >= 0.75) & (f <= 0.95)) | ((f > 1.05) & (f <= 2.5)))
    frange = np.take(f, fmask)
    HR = np.take(frange, np.argmax(np.take(pxx, fmask), 0))[0] * 60
    return HR, np.take(f, fmask), np.take(pxx, fmask)

def audio_sync(win_sound_file,video_sound_file):
    
    fig1 = np.memmap(win_sound_file, dtype='h', mode='r')
    fig2 = np.memmap(video_sound_file, dtype='h', mode='r')
    
    fig1 = np.array(list(fig1)).flatten()
    fig2 = np.array(list(fig2)).flatten()
    print(fig1,fig2)
    fig_mask1 = np.argwhere((fig1>=-30000)&(fig1<=30000))
    fig1 =np.take(fig1,fig_mask1)
    fig_mask2 = np.argwhere((fig2>=-30000)&(fig2<=30000))
    fig2 =np.take(fig2,fig_mask2)
    scaler = MinMaxScaler(feature_range=(0, 1))  #将数据归一到0到1，可以根据数据特点归一到-1到1
    fig1 = scaler.fit_transform(fig1)  #归一化
    fig2 = scaler.fit_transform(fig2)  #归一化
    fig1 = np.array(list(fig1)).flatten()
    fig2 = np.array(list(fig2)).flatten()
    fig1[0:5]=0
    fig2[0:5]=0
    fig1_corr,fig2_corr,delay = cross_corr(fig1,fig2)

    delay_time = delay/16000
    return delay_time

def ROI2XY(roi):
    """
    transfer [[x,y,h,w]] to [x1:x2, y1:y2]
    return y1, y2, x1, x2
    """ 
    x = roi[0,0]
    y = roi[0,1]
    w = roi[0,2]
    h = roi[0,3]
    
    return y, y+h, x, x+w 

def next_power_of_2(x):
    return 1 if x == 0 else 2**(x - 1).bit_length()

## Load RGB
if __name__ == "__main__":
    
    origin_path =  r'C:\Users\T JACK\Desktop\ACSP\code\GreenChannel\dataset'#数据集根目录
    current_path = os.path.join(origin_path,str(subject_id))#实验者目录
    os.chdir(current_path) #切换当前路径

    #ROI
    ROI_path = "ROI.mat" #---------------TBD
    ROI_dict = sio.loadmat(ROI_path)
    roi_names = ["Forehead", "left", "right", "curtain"]

    #excel
    excel_path = str(subject_id)+'.xlsx'
    wb = load_workbook(excel_path)
    ws = wb['实验信息总表']
    ws.cell(1,21).value = 'gt_HR'
    ws.cell(1,5).value = 'lasting_time'
    ws.cell(1,22).value = 'delay_time(ms)'
    for i in range(3):
        ws.cell(1,6+5*i).value = 'peak_delay_'+roi_names[i]
        ws.cell(1,7+5*i).value = 'cross_delay_'+roi_names[i]
        ws.cell(1,8+5*i).value = 'MAE_'+roi_names[i]
        ws.cell(1,9+5*i).value = 'SNR_'+roi_names[i]
        ws.cell(1,10+5*i).value = 'Face_HR_'+roi_names[i]

    for i in range(experiment_nums):
        experiment_id = i
        
        filepath = str(subject_id)+'_'+str(experiment_id)+'.mp4' 
        cut_filepath = str(subject_id)+'_'+str(experiment_id)+'_cut.mp4' 
        wav_file = str(subject_id)+'_'+str(experiment_id)+'.wav'
        wavefile1 = str(subject_id)+'_'+str(experiment_id)+'start.wav'
        wavefile2 = str(subject_id)+'_'+str(experiment_id)+'end.wav'
        extract_audio(filepath,wav_file)
        delay_time_1 = audio_sync(wavefile1,wav_file)
        delay_time_2 = audio_sync(wavefile2,wav_file)
        print(delay_time_1,delay_time_2)
        start_win_record_time = ws.cell(experiment_id+2,4).value
        record_delay_time = ws.cell(experiment_id+2,5).value
        record_delay_time = int(abs(record_delay_time-(delay_time_2-delay_time_1))*1000)
        if record_delay_time >10:
            print("同步异常 : ", record_delay_time)
        ws.cell(experiment_id+2,5).value = record_delay_time
        # print('start_win_record_time: ',start_win_record_time)

        

        #---------------TBD
        
        # RGB = skvideo.io.vread(filepath) #skvideo读取视频文件

        RGB = read_video(filepath)
        video_duration_time = 60
        # extract_video(filepath,cut_filepath,seconds_to_time(delay_time_1))
        # RGB_cut = skvideo.io.vread(cut_filepath)
        # print('after cut length : ', RGB_cut.shape[0])
        print('audio delay time:',delay_time_1)

    #----------------------------Load GT, RGB, roi----------------------------
        GT_path = excel_path #---------------TBD
        
        print("working on: ", filepath)
        
        #GT
        gt_pulse, gt_time = generate_pulse_gt(GT_path,experiment_id)
        gt_pulse = np.array(gt_pulse)

        
    
        Forehead = ROI_dict.get("Forehead")
        left = ROI_dict.get("left")
        right = ROI_dict.get("right")
        curtain = ROI_dict.get("curtain")
        # curtain2 = ROI_dict.get("curtain2")
        
        # ROIs = [Forehead, left, right, curtain1, curtain2]
        ROIs = [Forehead, left, right, curtain]
        
        # roi_names = ["Forehead", "left"]

        fps=60
        # video_duration_time = 50
        
        # gap = 1000 // fps
        print("fps: ", fps)
        
        # Result_dict = {}
        
        #Process GT
        gt_pulse = (gt_pulse-np.min(gt_pulse))
        gt_pulse = gt_pulse / np.max(gt_pulse)
        gt_start_time = int(gt_time[0])
        gt_length = len(gt_time)
        # gt_end_time = gt_time[gt_length-1]
        # time_duration = (gt_end_time-gt_start_time)/1000
        # start_time = 0
        gt_end_time = gt_start_time + video_duration_time * 1000
        
        
        # Result_dict["GT_HR"] = gt_HR
        # Result_dict["GT_f"] = gt_f
        # Result_dict["GT_pxx"] = gt_pxx

        #截取视频长度
        
        RGB_length = RGB.shape[0]
        print(RGB_length)
        # output_file = 'audio.wav'
        # extract_audio(filepath,output_file)
        # start_time,end_time = get_wav_duration(output_file)
        full_video_duration =video_duration(filepath)
        start_time = delay_time_1+ (gt_start_time-start_win_record_time)/1000
        end_time = start_time + video_duration_time
        # end_time = start_time+time_duration
        start_frame = int(start_time*RGB_length/full_video_duration)
        end_frame = int(end_time*RGB_length/full_video_duration)
        # end_frame = start_frame + gt_length
        RGB = RGB[start_frame:end_frame]
        # video_duration_time = video_duration(filepath)
        
        # print("可用得到视频长度为： ",video_duration_time)
        RGB_length = RGB.shape[0]
        print(RGB_length)
        for i in range(len(gt_pulse)):
            if gt_time[i]>=gt_end_time:
                print(i,gt_time[i])
                gt_pulse = gt_pulse[0:i]
                gt_time = gt_time[0:i]
                break
                
        re_time = np.linspace(gt_start_time,gt_end_time,RGB_length)
        re_gt_pulse = np.interp(re_time, gt_time, gt_pulse)
        gt_length = len(re_gt_pulse)
        # max_abs_scaler = preprocessing.MaxAbsScaler()
        # re_gt_pulse = max_abs_scaler.fit_transform(re_gt_pulse)
        gt_HR, gt_f, gt_pxx = calculate_HR(fps, re_gt_pulse)
        print("GT HR: ", gt_HR)
        

        #----------------------------Crop RGB & Save mae + SNR + pxx----------------------------
        MAE_sum = []
        SNR_sum = []
        HR_sum = []
        cross_delay_sum =[]
        peak_delay_sum = []
        bottom_delay_sum = []
        plt.clf()
        fig = plt.figure(figsize=(30,9))

        for i, roi in enumerate(ROIs):
        # for i in range(5):
            roi = ROIs[i]
            y1, y2, x1, x2 = ROI2XY(roi)
            # print(y1,y2,x1,x2)
            Crop_G = RGB[:, y1:y2, x1:x2,1]
                
            # plt.imsave(filepath + "img_crop1stFrame/" + file[:-4] + "_" + roi_names + '.png', crop_img)
            
            #Average
            Crop_G_mean = np.mean(Crop_G, axis=(1,2))
            # print("mean shape: ", Crop_G_mean)
            Crop_G_mean_diff = np.diff(Crop_G_mean)
            
            pulse_pred = detrend(Crop_G_mean, 100)
            [b_pulse, a_pulse] = butter(1, [0.7 / fps * 2, 2.6 / fps * 2], btype='bandpass')
            Crop_G_mean = scipy.signal.filtfilt(b_pulse, a_pulse, np.double(pulse_pred))
            Crop_G_mean = (Crop_G_mean-np.min(Crop_G_mean))
            Crop_G_mean = Crop_G_mean / np.max(Crop_G_mean)
            #FFT, save result as mat
            HR, f, pxx = calculate_HR(fps, Crop_G_mean)
            # Result_dict[roi_names[i] + "_HR"] = HR
            # Result_dict[roi_names[i] + "_f"] = f
            # Result_dict[roi_names[i] + "_pxx"] = pxx
            # #MAE & SNR
            HR = round(HR,3)
            gt_HR = round(gt_HR,3)
            MAE_sum.append(abs(HR - gt_HR))
            SNR = bvpsnr(Crop_G_mean, fps, HR, False)
            # Result_dict[roi_names[i] + "_SNR"] = SNR
            SNR = round(SNR,3)
            HR_sum.append(HR)
            SNR_sum.append(SNR)

            #before_FFT
            # plt.subplot(211)

            # ax1 = plt.subplot(2,1,1)
            # plt.plot(Crop_G_mean)

            # ax2 = plt.subplot(2,1,2)
            # plt.plot(re_gt_pulse)

            #使用自相关
            print(len(Crop_G_mean),len(re_gt_pulse))
            # print(np.correlate(Crop_G_mean,re_gt_pulse,mode='full'))
            s1,s2_0,cross_delay = cross_corr(Crop_G_mean,re_gt_pulse)
            print(' cross delay time :',cross_delay)
            # if cross_delay >3550:
            #     cross_delay = 3600-cross_delay
            peak_delay_time = peak_delay(Crop_G_mean,re_gt_pulse,gt_HR)
            print('peak delay time: ',peak_delay_time)
            bottom_delay_time = bottom_delay(Crop_G_mean,re_gt_pulse,gt_HR)
            print('bottom delay time: ',bottom_delay_time)
            cross_delay_sum.append(cross_delay)
            peak_delay_sum.append(peak_delay_time)
            bottom_delay_sum.append(bottom_delay_time)


            # plt.figure()
            # print(len(Crop_G_mean),len(re_gt_pulse))
            ax1 = plt.subplot(2,4,i+1)
            plt.title(roi_names[i]+" cross :"+str(cross_delay)+" peak: "+str(peak_delay_time ))
            plt.plot(np.reshape(Crop_G_mean,[-1,1]))
            plt.plot(np.reshape(re_gt_pulse,[-1,1]))
            plt.legend(['camera', 'oximeter'], loc='upper right')


            # #Save FFT pngs
            plt.subplot(2,4,i+5)
           
            plt.plot(f, np.reshape(pxx, [-1, 1]))
            plt.title(roi_names[i] + ' HR: ' + str(HR) + " GT: "+str(gt_HR)+" SNR: " + str(SNR))
        
        plt.show()
        # plt.savefig(str(subject_id)+'_'+str(experiment_id)+'_result.png', dpi=300) 
        plt.savefig(str(subject_id)+'_'+str(experiment_id)+'_result.png')#---------------TBD
        plt.close()
        
        # sio.savemat("after_green.mat", Result_dict) #---------------TBD
        # ws.cell(experiment_id+2,5).value = gt_HR #GT_HR
        for i in range(3):
            ws.cell(experiment_id+2,6+i*5).value =  peak_delay_sum[i]#peak_delay
            ws.cell(experiment_id+2,7+i*5).value =  cross_delay_sum[i] #cross_delay
            ws.cell(experiment_id+2,8+i*5).value =  MAE_sum[i]#MAE
            ws.cell(experiment_id+2,9+i*5).value =  SNR_sum[i]#SNR
            ws.cell(experiment_id+2,10+i*5).value = HR_sum[i] #Face_HR
        
    wb.save(excel_path)    


    #leftCheak

    #RightCheak

    #Curtain1

    #Curtain2


#------------------------------------------------------------------------------#
## Start cropping

# ## Select ROI
# # videodata = videodata[200:-100]
# sample_image = videodata[0]
# r = cv2.selectROI("face", sample_image) # Press enter after selecting box
# print('Coordiantes: ', r)
# imCrop = sample_image[int(r[1]):int(r[1]+r[3]), int(r[0]):int(r[0]+r[2])]
# cv2.imshow("Image", imCrop)
# cv2.waitKey(0) # Press enter again to close both windows
# cv2.destroyWindow("face")
# cv2.destroyWindow("Image")


# ## Region of interest
# print('Start processing ROI')
# length = videodata.shape[0]
# roi_data = []
# for i in range(length):
#     im = videodata[i]
#     imCrop = im[int(r[1]):int(r[1]+r[3]), int(r[0]):int(r[0]+r[2]), :]
#     roi_data.append(imCrop)
# roi_data = np.array(roi_data)
# print('roi shape', roi_data.shape)

# green_channel = roi_data[:,:, :, 1]
# green_channel_mean = green_channel.mean(axis=(1, 2))

# print("mean shape", green_channel_mean.shape)

# def next_power_of_2(x):
#     return 1 if x == 0 else 2**(x - 1).bit_length()

# fs = 30 ## Needs to be revised.
# green_channel_mean = np.diff(green_channel_mean)
# N = next_power_of_2(green_channel_mean.shape[0])
# f, pxx = scipy.signal.periodogram(green_channel_mean, fs=fs, nfft= N, detrend=False)
# fmask = np.argwhere((f >= 0.75) & (f <= 2.5))
# f, pxx  = np.take(f, fmask), np.take(pxx, fmask)
# plt.subplot(211) #
# plt.plot(f, np.reshape(pxx, [-1, 1]))
# plt.title('FFT')
# plt.subplot(212)
# plt.plot(green_channel_mean)
# plt.title('Green Channel')
# plt.show()





