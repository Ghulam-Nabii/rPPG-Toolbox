from email.mime import base
from preprocess import generate_pulse_gt
import serial
from flask import Flask
from flask import request, send_file, Flask, flash, request, redirect, url_for, Response
import os
import time
import csv
from werkzeug.utils import secure_filename
import requests
import serial
import winsound
import multiprocessing
import winsound
import sounddevice as sd
from scipy.io import wavfile
from multiprocessing.dummy import Pool as ThreadPool
import threading
from scipy.signal import chirp,spectrogram
import matplotlib.pyplot as plt
import numpy as np
import pyaudio
import wave 
from openpyxl import load_workbook
import pandas as pd  
import openpyxl
import cv2 

subject_id = 9
experiment_nums =20 #实验组数
excel_path = str(subject_id)+'.xlsx'
origin_path =  r'C:\Users\T JACK\Desktop\ACSP\code\GreenChannel\dataset'#数据集根目录
current_path = os.path.join(origin_path,str(subject_id))#实验者目录
os.chdir(current_path)
# experiment_df = pd.read_excel('0.xlsx',sheet_name='实验信息总表')
wb = load_workbook(excel_path)
# wb_data = openpyxl.Workbook()
ws_base = wb['实验信息总表']
# experiment_nums = len(experiment_df.index) #实验组数

name_str = []
for i in range (experiment_nums):
    # name_str.append(str(subject_id)+'_'+str(experiment_df['实验序号'][i])+'_'+'ISO'+str(experiment_df['ISO'][i])+'_'+'S'+str(experiment_df['Shutter'][i]))
    name_str.append(str(subject_id)+'_'+str(i))

global start_sound_play_time,start_win_record_time,end_win_record_time
timestamp_record = [] 
start_win_record_time=0
end_win_record_time=0
start_sound_play_time=0




def recording_sound(file_path):
    sec = 5 # s
    #创建对象
    pa = pyaudio.PyAudio()
    # with wave.open(file_path,'w') as wf:
    #创建流：采样位，声道数，采样频率，缓冲区大小，input
    stream = pa.open(format=pyaudio.paInt16,
                    channels=1,
                    rate=16000,
                    input=True,
                    input_device_index= 1,
                    frames_per_buffer=1024)
    #创建式打开音频文件
    wf = wave.open(file_path, "wb")
    #设置音频文件的属性：采样位，声道数，采样频率，缓冲区大小，必须与流的属性一致
    wf.setnchannels(1)
    wf.setsampwidth(pa.get_sample_size(pyaudio.paInt16))
    wf.setframerate(16000)
    print("开始录音")
    #采样频率*秒数/缓冲区大小
    print('start record:',time.time())
    start_win_record_time = time.time()
    for w in range(int(16000*sec/1024)):
        data = stream.read(1024)#每次从流的缓存中读出数据
        wf.writeframes(data)#把读出的数据写入wf
    print('end record:',time.time())
    end_win_record_time = time.time()

    print("录音结束")
    
    stream.stop_stream()#先把流停止
    stream.close()#再把流关闭
    pa.terminate()#把对象关闭
    wf.close()#把声音文件关闭，否则不能播放
    timestamp_record.append(start_win_record_time)
    timestamp_record.append(end_win_record_time)
    return file_path

def start_sound_wav_out():
    fs = 16000
    t = np.linspace(0,3,fs)
    w = chirp(t, f0=2000, f1=200, t1=3, method='linear')
    start_sound_play_time =time.time()
    sd.play(w,fs)
    print('start play :',start_sound_play_time)
    timestamp_record.append(start_sound_play_time)
    return start_sound_play_time

def end_sound_wav_out():
    fs = 16000
    t = np.linspace(0,3,fs)
    w = chirp(t, f0=300, f1=2500, t1=3, method='linear')
    start_sound_play_time =time.time()
    sd.play(w,fs)
    print('end play :',start_sound_play_time)
    timestamp_record.append(start_sound_play_time)
    return start_sound_play_time


def get_oximeter_data(start_file_path,end_file_path,experiment_order,ser,wb):
    # timestamp_record =[]
    # os.system('start dist/usbserial.exe')
    try:
        
    #     portx="COM3"
    # #波特率，标准值之一：50,75,110,134,150,200,300,600,1200,1800,2400,4800,9600,19200,38400,57600,115200
    #     bps=9600
    #     #超时设置,None：永远等待操作，0为立即返回请求结果，其他值为等待超时时间(单位为秒）
    #     timex=5
    #         # 打开串口，并得到串口对象
    #     ser=serial.Serial(portx,bps,timeout=timex)

        # wb = load_workbook(excel_path)
        wb.create_sheet(index = experiment_order+1,title = str(experiment_order))
        ws = wb[str(experiment_order)]
        print(ws)

        # ser.close()#关闭串口
        # csvFile = open('PPG_Wave.csv', 'w', newline='', encoding='utf-8')
        # csv_writer = csv.writer(csvFile)
        

        base_start_time = int(round(1000 * time.time()))
        
        task1 = threading.Thread(target=recording_sound,args=(start_file_path,))
        task2 = threading.Thread(target=start_sound_wav_out)
        task3 = threading.Thread(target=recording_sound,args=(end_file_path,))
        task4 = threading.Thread(target=end_sound_wav_out)

        #起始音波
        task1.start()
        time.sleep(1)
        task2.start()
        time.sleep(1)


        


#60000ms: oximeter time

        while int(round(1000 * time.time())) - base_start_time <= 65000:
            

            # print("串口详情参数：", ser)
            if  ser.read().hex() != 'ff' and ser.read().hex() != '00':
                h = ser.read().hex()
                print(h)
                relative_time = str(int(round(1000 * time.time())) - base_start_time)
            #读一个字节
                
                # csv_writer.writerow([relative_time, h])
                ws.append([relative_time, h])
        # winsound.Beep(2000,500)
        
        #结束音波
        task3.start()
        time.sleep(1)
        task4.start()
        time.sleep(5)

        print(timestamp_record)
        print('This turn is OVER!')
        start_win_record_time = int(round(1000 *timestamp_record[1]))
        
        end_win_record_time = int(round(1000 *timestamp_record[2]))
        start_sound_play_time = int(round(1000 *timestamp_record[0]))
        start_win_record_time -= base_start_time
        start_sound_play_time -= base_start_time
        end_win_record_time -= base_start_time

        print('start_win_record_time :',start_win_record_time)
        print('end_win_record_time :', end_win_record_time)
        print('start_sound_play_time :',start_sound_play_time)
        start_oximeter_time = ws.cell(1,1).value
        print('start_oximeter_time :',start_oximeter_time)
        record_delay = timestamp_record[4]-timestamp_record[1]
        # time.sleep(5)
        # wb.save(excel_path)
        # time.sleep(5)

        return start_win_record_time, record_delay

        
    except Exception as e:
        print("---异常--- ",e)
        return 'Hello, World!'


if __name__ == "__main__":
    
    portx="COM3"
#波特率，标准值之一：50,75,110,134,150,200,300,600,1200,1800,2400,4800,9600,19200,38400,57600,115200
    bps=9600
    #超时设置,None：永远等待操作，0为立即返回请求结果，其他值为等待超时时间(单位为秒）
    timex=5
        # 打开串口，并得到串口对象
    ser=serial.Serial(portx,bps,timeout=timex)
    print("串口详情参数：", ser)
    start_file = 'start.wav'
    end_file = 'end.wav'


    for i in range(experiment_nums):
        start_file_path = name_str[i]+start_file
        end_file_path = name_str[i]+end_file
        timestamp_record = []
        print('--------------',i+1,'--------------')
        os.system('pause')
        time.sleep(1)
        print(i,'start ',time.time())
        start_win_record_time, record_delay = get_oximeter_data(start_file_path,end_file_path,i,ser,wb)
        print(i,'end ',time.time())
        ws_base.cell(i+2,4).value = start_win_record_time
        ws_base.cell(i+2,5).value = record_delay
    wb.save(excel_path)