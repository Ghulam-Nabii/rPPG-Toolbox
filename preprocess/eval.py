
import matplotlib
matplotlib.use('Agg')#仅保存图片不显示

from flask import before_render_template
import skvideo.io
import numpy as np
import matplotlib.pyplot as plt
import scipy
import scipy.io as sio
import os
from utils import bottom_delay, bvpsnr, peak_delay,video_duration,cross_corr,read_video
from preprocess import generate_pulse_gt
from mp4_wav import *
from inference_preprocess import detrend
from scipy.signal import butter
from sklearn import preprocessing
from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook
HDF5_DISABLE_VERSION_CHECK = 1

#experiments---------------TBD
subject_id = 35
experiment_nums = 20

def calculate_HR(fs, signal):
    # N = 30 * fs
    N = 30 * fs
    pulse_fft = np.expand_dims(signal, 0)
    f, pxx = scipy.signal.periodogram(pulse_fft, fs=fs, nfft=4 * N, detrend=False)
    fmask = np.argwhere((f >= 0.75) & (f <= 2.5))  # regular Heart beat are 0.75*60 and 2.5*60
    # fmask = np.argwhere(((f >= 0.75) & (f <= 0.95)) | ((f > 1.05) & (f <= 2.5)))
    frange = np.take(f, fmask)
    HR = np.take(frange, np.argmax(np.take(pxx, fmask), 0))[0] * 60
    return HR, np.take(f, fmask), np.take(pxx, fmask)



def ROI2XY(roi):
    """
    transfer [[x,y,h,w]] to [x1:x2, y1:y2]
    return y1, y2, x1, x2
    """ 
    x = roi[0,0]
    y = roi[0,1]
    w = roi[0,2]
    h = roi[0,3]
    
    return y, y+h, x, x+w 

def next_power_of_2(x):
    return 1 if x == 0 else 2**(x - 1).bit_length()

## Load RGB
if __name__ == "__main__":
    
    origin_path =  r'E:\rPPG\dataset'#数据集根目录
    current_path = os.path.join(origin_path,str(subject_id))#实验者目录
    os.chdir(current_path) #切换当前路径

    #ROI
    ROI_path = "ROI.mat" #---------------TBD
    ROI_dict = sio.loadmat(ROI_path)
    roi_names = ["Forehead", "left", "right", "curtain"]

    #excel
    excel_path = str(subject_id)+'.xlsx'
    wb = load_workbook(excel_path)
    ws = wb['实验信息总表']
    ws.cell(1,21).value = 'gt_HR'
    ws.cell(1,5).value = 'lasting_time'
    ws.cell(1,22).value = 'delay_time(ms)'
    for i in range(3):
        ws.cell(1,6+5*i).value = 'peak_delay_'+roi_names[i]
        ws.cell(1,7+5*i).value = 'cross_delay_'+roi_names[i]
        ws.cell(1,8+5*i).value = 'MAE_'+roi_names[i]
        ws.cell(1,9+5*i).value = 'SNR_'+roi_names[i]
        ws.cell(1,10+5*i).value = 'Face_HR_'+roi_names[i]

    for i in range(experiment_nums):
        experiment_id = i
        prefix ='E:/rPPG/processed_dataset/'+ 'subject'+str(subject_id)+'_'+str(experiment_id)
        filepath = prefix+'/video.mp4' 
        GT_path = prefix +'/bvp.npy'
        #---------------TBD
        

    #----------------------------Load GT, RGB, roi----------------------------
        #RGB
        RGB = read_video(filepath)
        print("working on: ", filepath)
        
        #GT
        gt_pulse = np.load(GT_path)
        
        Forehead = ROI_dict.get("Forehead")
        left = ROI_dict.get("left")
        right = ROI_dict.get("right")
        curtain = ROI_dict.get("curtain")
        # curtain2 = ROI_dict.get("curtain2")
        
        # ROIs = [Forehead, left, right, curtain1, curtain2]
        ROIs = [Forehead, left, right, curtain]
        
        # roi_names = ["Forehead", "left"]

        fps=30.0

        
        # gap = 1000 // fps
        print("fps: ", fps)
        
        #Process GT
        gt_pulse = (gt_pulse-np.min(gt_pulse))
        gt_pulse = gt_pulse / np.max(gt_pulse)


        RGB_length = RGB.shape[0]          
        gt_HR, gt_f, gt_pxx = calculate_HR(fps,gt_pulse)
        gt_HR = round(gt_HR,3)
        print("GT HR: ", gt_HR)
        

        #----------------------------Crop RGB & Save mae + SNR + pxx----------------------------
        MAE_sum = []
        SNR_sum = []
        HR_sum = []
        cross_delay_sum =[]
        peak_delay_sum = []
        bottom_delay_sum = []
        plt.clf()
        fig = plt.figure(figsize=(25,9))

        for i, roi in enumerate(ROIs):
        # for i in range(5):
            roi = ROIs[i]
            y1, y2, x1, x2 = ROI2XY(roi)
            # print(y1,y2,x1,x2)
            Crop_G = RGB[:, y1:y2, x1:x2,1]
            
            #Average
            Crop_G_mean = np.mean(Crop_G, axis=(1,2))
            # print("mean shape: ", Crop_G_mean)
            Crop_G_mean_diff = np.diff(Crop_G_mean)
            
            pulse_pred = detrend(Crop_G_mean, 100)
            [b_pulse, a_pulse] = butter(1, [0.7 / fps * 2, 2.6 / fps * 2], btype='bandpass')
            Crop_G_mean = scipy.signal.filtfilt(b_pulse, a_pulse, np.double(pulse_pred))
            Crop_G_mean = (Crop_G_mean-np.min(Crop_G_mean))
            Crop_G_mean = Crop_G_mean / np.max(Crop_G_mean)
            #FFT, save result as mat
            HR, f, pxx = calculate_HR(fps, Crop_G_mean)
            # #MAE & SNR
            MAE_sum.append(abs(HR - gt_HR))
            SNR = bvpsnr(Crop_G_mean, fps, HR, False)
            # Result_dict[roi_names[i] + "_SNR"] = SNR
            SNR = round(SNR,3)
            HR= round(HR,3)
            HR_sum.append(HR)
            SNR_sum.append(SNR)

            #before_FFT
            # plt.subplot(211)

            # ax1 = plt.subplot(2,1,1)
            # plt.plot(Crop_G_mean)

            # ax2 = plt.subplot(2,1,2)
            # plt.plot(re_gt_pulse)

            #使用自相关
            print(len(Crop_G_mean),len(gt_pulse))
            # print(np.correlate(Crop_G_mean,re_gt_pulse,mode='full'))
            s1,s2_0,cross_delay = cross_corr(Crop_G_mean,gt_pulse)
            print('cross delay time :',cross_delay)
            # if cross_delay >3550:
            #     cross_delay = 3600-cross_delay
            peak_delay_time = peak_delay(Crop_G_mean,gt_pulse,gt_HR)
            print('peak delay time: ',peak_delay_time)
            bottom_delay_time = bottom_delay(Crop_G_mean,gt_pulse,gt_HR)
            print('bottom delay time: ',bottom_delay_time)
            cross_delay_sum.append(cross_delay)
            peak_delay_sum.append(peak_delay_time)
            bottom_delay_sum.append(bottom_delay_time)


            # plt.figure()
            # print(len(Crop_G_mean),len(re_gt_pulse))
            ax1 = plt.subplot(2,4,i+1)
            plt.title(roi_names[i]+" cross :"+str(cross_delay)+" peak: "+str(peak_delay_time ))
            plt.plot(np.reshape(Crop_G_mean,[-1,1]))
            plt.plot(np.reshape(gt_pulse,[-1,1]))
            plt.legend(['camera', 'oximeter'], loc='upper right')


            # #Save FFT pngs
            plt.subplot(2,4,i+5)
           
            plt.plot(f, np.reshape(pxx, [-1, 1]))
            plt.title(roi_names[i] + ' HR: ' + str(HR) + " GT: "+str(gt_HR)+" SNR: " + str(SNR))
        
        plt.show()
        # plt.savefig(str(subject_id)+'_'+str(experiment_id)+'_result.png', dpi=300) 
        plt.savefig(str(subject_id)+'_'+str(experiment_id)+'_result.png')#---------------TBD
        plt.close()
        
        # sio.savemat("after_green.mat", Result_dict) #---------------TBD
        ws.cell(experiment_id+2,21).value = gt_HR #GT_HR
        for i in range(3):
            ws.cell(experiment_id+2,6+i*5).value =  peak_delay_sum[i]#peak_delay
            ws.cell(experiment_id+2,7+i*5).value =  cross_delay_sum[i] #cross_delay
            ws.cell(experiment_id+2,8+i*5).value =  MAE_sum[i]#MAE
            ws.cell(experiment_id+2,9+i*5).value =  SNR_sum[i]#SNR
            ws.cell(experiment_id+2,10+i*5).value = HR_sum[i] #Face_HR
        
    wb.save(excel_path)    










